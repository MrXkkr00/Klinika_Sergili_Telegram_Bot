import openpyxl
from aiogram import types
from aiogram.dispatcher import FSMContext
from aiogram.dispatcher.filters.builtin import CommandStart
from aiogram.dispatcher.filters.state import StatesGroup, State
from aiogram.types import ReplyKeyboardRemove
# from utils.misc.get_distance import choose_shortest

from keyboards.default.bosh_menu import menu, keyboard_contact
from loader import dp, bot

wb = openpyxl.load_workbook('data/Dorixona.xlsx')
ws = wb.active
values_1 = [ws.cell(row=i, column=2).value for i in range(1, ws.max_row + 1)]
values_2 = [ws.cell(row=i, column=3).value for i in range(1, ws.max_row + 1)]
values_narx = [ws.cell(row=i, column=4).value for i in range(1, ws.max_row + 1)]

maxrow = ws.max_row + 1 - 2
values_0 = "iphone"


class StateNarx(StatesGroup):
    narx1 = State()
    narx2 = State()


@dp.message_handler(text="🏷 Narxlar")
@dp.message_handler(text="🏷 Цены")
async def bt_start(message: types.Message):
    global values_0
    await message.answer(f"1. {values_1[1]}\n")
    for i in range(2, maxrow):
        # seria == values_series[i] and model == values_phone[i] and memory == values_memory[i] and color == \
        # values_color[i] and
        if values_1[i-1] != values_1[i]:
            # and model == values_phone[i] and memory == values_memory[i] and tel_color != values_color[i]:

            await message.answer(f"{i}. {values_1[i]}\n")

    await message.answer(f"Kerak xizmat raqamini yuboring\nПожалуйста, отправьте требуемый сервисный номер", reply_markup=ReplyKeyboardRemove())
    await StateNarx.narx1.set()


@dp.message_handler(content_types='location')
async def get_contact(message: types.Message):
    location = message.location
    latitude = location.latitude
    longitude = location.longitude

    await message.answer(f"Rahmat. \n"
                         f"Latitude = {latitude}\n"
                         f"Longitude = {longitude}\n\n")




@dp.message_handler(lambda message: not message.text.isdigit(), state=StateNarx.narx1)
async def process_age_ild(message: types.Message):
    """
    If age is invalid
    """
    return await message.reply("Bu yerga faqat raqam kiriting( xizmat raqami ):\nMassalan: 125")


@dp.message_handler(lambda message: message.text.isdigit(), state=StateNarx.narx1)
async def answer_narx1(message: types.Message, state: FSMContext):
    global values_0
    narx11 = int(message.text)
    await state.update_data(
        {"narx11": narx11})
    for i in range(1, maxrow):
        # seria == values_series[i] and model == values_phone[i] and memory == values_memory[i] and color == \
        # values_color[i] and
        if values_1[narx11] == values_1[i] and values_0 != values_2[i]:
            # and model == values_phone[i] and memory == values_memory[i] and tel_color != values_color[i]:
            values_0 = values_2[i]
            await message.answer(f"{values_2[i]}\n\n💵Narx:  {values_narx[i]}", reply_markup=menu)
    await state.finish()
